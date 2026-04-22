#!/usr/bin/env python3
“””
excel_to_md.py
Excel方眼紙・管理会計PLなどをMarkdown/JSONに変換するツール

対応パターン:

- セル結合による見出し・レイアウト
- 罫線によるテーブル領域の検出
- 列オフセット（空白セル）による字下げ階層
- 矢印・記号の保持/変換
- 複数シート対応
- 空行・罫線によるセクション分割

使い方:
python excel_to_md.py input.xlsx [output.md]
python excel_to_md.py input.xlsx –format json [output.json]
python excel_to_md.py input.xlsx –sheet “Sheet1” –format md
“””

import sys
import json
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Any

try:
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
except ImportError:
print(“openpyxl が必要です: pip install openpyxl”, file=sys.stderr)
sys.exit(1)

# ──────────────────────────────────────────────

# データ構造

# ──────────────────────────────────────────────

@dataclass
class Cell:
value: Any = None
is_merged_main: bool = False   # 結合セルの左上（代表）
is_merged_sub: bool = False    # 結合セルの非代表（空扱い）
merge_span_col: int = 1        # 横方向の結合幅
merge_span_row: int = 1        # 縦方向の結合幅
border_top: bool = False
border_bottom: bool = False
border_left: bool = False
border_right: bool = False
is_bold: bool = False
bg_color: Optional[str] = None
indent: int = 0                # openpyxl alignment indent

```
@property
def text(self) -> str:
    if self.value is None:
        return ""
    return str(self.value).strip()

@property
def is_empty(self) -> bool:
    return self.text == ""

@property
def has_any_border(self) -> bool:
    return any([self.border_top, self.border_bottom, self.border_left, self.border_right])

@property
def has_full_border(self) -> bool:
    return all([self.border_top, self.border_bottom, self.border_left, self.border_right])
```

@dataclass
class GridRow:
cells: list[Cell]

```
def is_empty(self) -> bool:
    return all(c.is_empty or c.is_merged_sub for c in self.cells)

def non_empty_cells(self) -> list[tuple[int, Cell]]:
    return [(i, c) for i, c in enumerate(self.cells) if not c.is_empty and not c.is_merged_sub]

def first_nonempty_col(self) -> int:
    for i, c in enumerate(self.cells):
        if not c.is_empty and not c.is_merged_sub:
            return i
    return 0
```

@dataclass
class Section:
kind: str          # “table” | “header” | “freeform” | “separator”
rows: list[GridRow] = field(default_factory=list)
title: str = “”
indent_base: int = 0

# ──────────────────────────────────────────────

# グリッド構築

# ──────────────────────────────────────────────

def _border_has_line(border_side) -> bool:
“”“openpyxl border side が実線かどうか”””
if border_side is None:
return False
style = border_side.border_style
return style is not None and style not in (“none”, “hair”)

def build_grid(sheet) -> list[GridRow]:
“”“シートをCell[][]グリッドに変換”””
max_row = sheet.max_row or 1
max_col = sheet.max_column or 1

```
# 結合セル情報を収集
# key: (row, col) 0-indexed → (span_row, span_col, is_main)
merge_info: dict[tuple[int, int], tuple[int, int, bool]] = {}
for merge_range in sheet.merged_cells.ranges:
    min_r, min_c = merge_range.min_row - 1, merge_range.min_col - 1
    max_r, max_c = merge_range.max_row - 1, merge_range.max_col - 1
    span_r = max_r - min_r + 1
    span_c = max_c - min_c + 1
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            is_main = (r == min_r and c == min_c)
            merge_info[(r, c)] = (span_r, span_c, is_main)

grid: list[GridRow] = []
for r in range(1, max_row + 1):
    cells = []
    for c in range(1, max_col + 1):
        ws_cell = sheet.cell(row=r, column=c)
        mi = merge_info.get((r - 1, c - 1))

        cell = Cell()
        cell.value = ws_cell.value

        if mi:
            span_r, span_c, is_main = mi
            if is_main:
                cell.is_merged_main = True
                cell.merge_span_row = span_r
                cell.merge_span_col = span_c
            else:
                cell.is_merged_sub = True
                cell.value = None  # 非代表は空扱い

        # 罫線
        border = ws_cell.border
        cell.border_top = _border_has_line(border.top)
        cell.border_bottom = _border_has_line(border.bottom)
        cell.border_left = _border_has_line(border.left)
        cell.border_right = _border_has_line(border.right)

        # フォント
        if ws_cell.font:
            cell.is_bold = bool(ws_cell.font.bold)

        # 背景色
        if ws_cell.fill and ws_cell.fill.fgColor:
            raw = ws_cell.fill.fgColor.rgb
            if raw and raw not in ("00000000", "FFFFFFFF", "FF000000"):
                cell.bg_color = raw

        # セル内インデント
        if ws_cell.alignment and ws_cell.alignment.indent:
            cell.indent = int(ws_cell.alignment.indent)

        cells.append(cell)
    grid.append(GridRow(cells=cells))

return grid
```

# ──────────────────────────────────────────────

# テーブル・セクション検出

# ──────────────────────────────────────────────

def detect_col_count(grid: list[GridRow]) -> int:
if not grid:
return 0
return max(len(row.cells) for row in grid)

def detect_indent_level(row: GridRow, base_col: int = 0) -> int:
“””
行の階層レベルを推定する。
- セル内alignmentインデント
- 先頭空白セル数（方眼紙スタイル）
を組み合わせて判定。
“””
nonempty = row.non_empty_cells()
if not nonempty:
return 0
first_col, first_cell = nonempty[0]
# セル内インデントが明示されていればそれを使う
if first_cell.indent > 0:
return first_cell.indent
# 列オフセットによる字下げ
return max(0, first_col - base_col)

def _is_section_header_row(row: GridRow, total_cols: int) -> bool:
“””
行全体がセクションヘッダーかどうかを推定:
- 結合セルが横幅の大部分を占める
- 太字
- 背景色あり
“””
nonempty = row.non_empty_cells()
if not nonempty:
return False
if len(nonempty) == 1:
_, cell = nonempty[0]
if cell.is_merged_main and cell.merge_span_col >= max(2, total_cols // 2):
return True
if cell.is_bold and total_cols > 2:
return True
return False

def _row_is_separator(row: GridRow) -> bool:
“”“空行か罫線のみの行”””
if row.is_empty():
return True
texts = [c.text for c in row.cells if not c.is_merged_sub]
return all(t == “” for t in texts)

def split_into_sections(grid: list[GridRow]) -> list[Section]:
“””
グリッドを意味のあるセクションに分割する。
セクション境界の判定:
1. 空行 2行以上
2. ヘッダー行（横幅結合+太字）
“””
total_cols = detect_col_count(grid)
sections: list[Section] = []
current_rows: list[GridRow] = []
current_title = “”

```
def flush(kind: str = "table"):
    nonlocal current_rows, current_title
    if current_rows:
        # indentの基準列を検出
        min_col = min(
            (r.first_nonempty_col() for r in current_rows if not r.is_empty()),
            default=0
        )
        s = Section(kind=kind, rows=current_rows[:], title=current_title, indent_base=min_col)
        sections.append(s)
        current_rows = []
        current_title = ""

consecutive_empty = 0

for row in grid:
    if _row_is_separator(row):
        consecutive_empty += 1
        current_rows.append(row)
        continue

    if consecutive_empty >= 2:
        flush("table")
        consecutive_empty = 0

    consecutive_empty = 0

    if _is_section_header_row(row, total_cols):
        flush("table")
        nonempty = row.non_empty_cells()
        if nonempty:
            current_title = nonempty[0][1].text
        continue

    current_rows.append(row)

flush("table")

# 空行のみのセクションを除去
return [s for s in sections if any(not r.is_empty() for r in s.rows)]
```

# ──────────────────────────────────────────────

# テーブル構造解析

# ──────────────────────────────────────────────

def trim_grid_rows(rows: list[GridRow]) -> list[GridRow]:
“”“先頭・末尾の空行を除去”””
i_start = 0
while i_start < len(rows) and rows[i_start].is_empty():
i_start += 1
i_end = len(rows) - 1
while i_end >= 0 and rows[i_end].is_empty():
i_end -= 1
return rows[i_start:i_end + 1]

def detect_table_header_rows(rows: list[GridRow], total_cols: int) -> int:
“””
テーブルのヘッダー行数を推定。
- 値が全て文字列の行
- 罫線の下罫線
- 背景色・太字
“””
if not rows:
return 1
# 1行目がヘッダーかどうか
first = rows[0]
nonempty = first.non_empty_cells()
if not nonempty:
return 1

```
# 数値が含まれているかどうか
has_number = any(
    isinstance(c.value, (int, float)) for _, c in nonempty
)
if not has_number:
    return 1
return 0
```

def rows_to_table_data(rows: list[GridRow], indent_base: int = 0) -> list[list[str]]:
“””
GridRow のリストを文字列2D配列に変換。
- 結合セルは代表セルにテキスト、残りは””
- 列オフセットインデントはテキストに反映
“””
if not rows:
return []
total_cols = max(len(r.cells) for r in rows)

```
result = []
for row in rows:
    row_data = []
    for c_idx in range(total_cols):
        if c_idx < len(row.cells):
            cell = row.cells[c_idx]
            if cell.is_merged_sub:
                row_data.append("")
            else:
                text = cell.text
                # セル内インデントをスペースで表現
                if cell.indent > 0:
                    text = "\u3000" * cell.indent + text
                row_data.append(text)
        else:
            row_data.append("")
    result.append(row_data)
return result
```

# ──────────────────────────────────────────────

# PL / 階層リスト検出

# ──────────────────────────────────────────────

def is_pl_like(rows: list[GridRow]) -> bool:
“””
管理会計PLっぽい構造かどうかを推定:
- 左列に科目名（文字列）
- 右列に数値
- インデントによる階層
“””
if len(rows) < 3:
return False
numeric_rows = 0
for row in rows:
nonempty = row.non_empty_cells()
if len(nonempty) >= 2:
# 最後の非空セルが数値
_, last = nonempty[-1]
if isinstance(last.value, (int, float)):
numeric_rows += 1
return numeric_rows >= len(rows) * 0.4

def rows_to_pl_markdown(rows: list[GridRow], indent_base: int = 0) -> str:
“””
PL形式のデータをMarkdownのリスト or テーブルに変換。
インデントを階層として表現する。
“””
total_cols = detect_col_count_from_rows(rows)
lines = []

```
# ヘッダー行（数値が入っていない行）を探す
header_rows = []
data_rows = []
header_done = False
for row in rows:
    if row.is_empty():
        continue
    nonempty = row.non_empty_cells()
    has_num = any(isinstance(c.value, (int, float)) for _, c in nonempty)
    if not header_done and not has_num:
        header_rows.append(row)
    else:
        header_done = True
        data_rows.append(row)

# ヘッダーがあればテーブルヘッダーとして使う
if header_rows and data_rows:
    hdata = rows_to_table_data(header_rows, indent_base)
    ddata = rows_to_table_data(data_rows, indent_base)
    all_data = hdata + ddata
    return _render_markdown_table(all_data, len(hdata))
else:
    # 全部テーブルとして出力
    all_data = rows_to_table_data(rows, indent_base)
    return _render_markdown_table(all_data, detect_table_header_rows(rows, total_cols))
```

def detect_col_count_from_rows(rows: list[GridRow]) -> int:
if not rows:
return 0
return max(len(r.cells) for r in rows)

# ──────────────────────────────────────────────

# Markdown レンダリング

# ──────────────────────────────────────────────

def _col_widths(data: list[list[str]]) -> list[int]:
if not data:
return []
ncols = max(len(r) for r in data)
widths = [0] * ncols
for row in data:
for i, cell in enumerate(row):
# 日本語文字は幅2でカウント
w = sum(2 if ord(c) > 0x7F else 1 for c in cell)
widths[i] = max(widths[i], w)
return [max(w, 3) for w in widths]

def _pad(text: str, width: int) -> str:
w = sum(2 if ord(c) > 0x7F else 1 for c in text)
return text + “ “ * max(0, width - w)

def _render_markdown_table(data: list[list[str]], header_row_count: int = 1) -> str:
“”“2D文字列配列をMarkdownテーブルとして描画”””
# 空の行を除去
data = [r for r in data if any(c.strip() for c in r)]
if not data:
return “”

```
ncols = max(len(r) for r in data)
# 列数を揃える
data = [r + [""] * (ncols - len(r)) for r in data]

# 末尾の空列を削除
while ncols > 1 and all(r[ncols - 1] == "" for r in data):
    ncols -= 1
    data = [r[:ncols] for r in data]

widths = _col_widths(data)
lines = []

for i, row in enumerate(data):
    line = "| " + " | ".join(_pad(row[j] if j < len(row) else "", widths[j]) for j in range(ncols)) + " |"
    lines.append(line)
    # ヘッダー区切り線
    if i == header_row_count - 1:
        sep = "| " + " | ".join("-" * widths[j] for j in range(ncols)) + " |"
        lines.append(sep)

return "\n".join(lines)
```

def _render_hierarchy_list(rows: list[GridRow], indent_base: int) -> str:
“””
インデントが明確な場合はMarkdownの箇条書き階層として出力。
数値列がある場合はインライン表示。
“””
lines = []
for row in rows:
if row.is_empty():
lines.append(””)
continue
level = max(0, detect_indent_level(row, indent_base))
prefix = “  “ * level + “- “

```
    nonempty = row.non_empty_cells()
    if not nonempty:
        continue

    # テキスト部分と数値部分を分離
    text_parts = []
    num_parts = []
    for _, cell in nonempty:
        if isinstance(cell.value, (int, float)):
            num_parts.append(cell.text)
        else:
            text_parts.append(cell.text)

    label = " ".join(text_parts)
    if num_parts:
        label += "  `" + "  /  ".join(num_parts) + "`"
    lines.append(prefix + label)

return "\n".join(l for l in lines if l.strip() or l == "")
```

# ──────────────────────────────────────────────

# 図形検出（罫線ベース）

# ──────────────────────────────────────────────

def detect_diagram_region(rows: list[GridRow]) -> bool:
“””
フローチャート・図形領域かどうかを判定する。
主な判定基準:
- 矢印文字（→↓など）が含まれる
- 数値セルがほぼなく、値が疎らに配置されている
- テーブル構造（行ヘッダー＋数値列）でない
数値データが多い場合はテーブル/PLとして扱い、図形判定しない。
“””
arrow_chars = {“→”, “←”, “↑”, “↓”, “⇒”, “⇔”, “►”, “▷”, “▼”, “△”, “▲”, “↗”, “↘”}
has_arrow = False
numeric_count = 0
total_nonempty = 0

```
for row in rows:
    for _, cell in row.non_empty_cells():
        total_nonempty += 1
        if any(a in cell.text for a in arrow_chars):
            has_arrow = True
        if isinstance(cell.value, (int, float)):
            numeric_count += 1

if total_nonempty == 0:
    return False

# 数値が多い場合はテーブルとして扱う
numeric_ratio = numeric_count / total_nonempty
if numeric_ratio > 0.25:
    return False

return has_arrow
```

def rows_to_code_block(rows: list[GridRow]) -> str:
“””
図形・方眼紙領域をコードブロック（固定幅）で表現。
罫線情報は `+`, `-`, `|` に変換しない（複雑すぎる）が、
セルの位置関係を空白で再現する。
“””
if not rows:
return “”
total_cols = detect_col_count_from_rows(rows)

```
# 列幅を決定
col_w = [0] * total_cols
for row in rows:
    for i, cell in enumerate(row.cells):
        if i < total_cols:
            w = sum(2 if ord(c) > 0x7F else 1 for c in cell.text)
            col_w[i] = max(col_w[i], w, 2)

lines = []
for row in rows:
    parts = []
    for i in range(total_cols):
        if i < len(row.cells):
            text = row.cells[i].text if not row.cells[i].is_merged_sub else ""
        else:
            text = ""
        w = sum(2 if ord(c) > 0x7F else 1 for c in text)
        parts.append(text + " " * max(0, col_w[i] - w + 1))
    lines.append("".join(parts).rstrip())

# 空行を除去（前後のみ）
while lines and not lines[0].strip():
    lines.pop(0)
while lines and not lines[-1].strip():
    lines.pop()

return "```\n" + "\n".join(lines) + "\n```"
```

# ──────────────────────────────────────────────

# セクション → Markdown変換

# ──────────────────────────────────────────────

def section_to_markdown(section: Section, heading_level: int = 2) -> str:
rows = trim_grid_rows(section.rows)
if not rows:
return “”

```
parts = []

# セクションタイトル
if section.title:
    parts.append("#" * heading_level + " " + section.title)
    parts.append("")

# PL/階層リスト検出（数値データがある場合は図形より優先）
if is_pl_like(rows):
    md = rows_to_pl_markdown(rows, section.indent_base)
    parts.append(md)
    return "\n".join(parts)

# 図形検出（矢印などがある場合）
if detect_diagram_region(rows):
    parts.append(rows_to_code_block(rows))
    return "\n".join(parts)

# インデントが大きい場合は階層リスト
max_indent = max(detect_indent_level(r, section.indent_base) for r in rows if not r.is_empty())
if max_indent >= 2:
    parts.append(_render_hierarchy_list(rows, section.indent_base))
    return "\n".join(parts)

# 通常テーブル
data = rows_to_table_data(rows, section.indent_base)
header_count = detect_table_header_rows(rows, detect_col_count_from_rows(rows))
md = _render_markdown_table(data, header_count)
if md:
    parts.append(md)

return "\n".join(parts)
```

# ──────────────────────────────────────────────

# JSON出力

# ──────────────────────────────────────────────

def grid_to_json(grid: list[GridRow]) -> dict:
“”“グリッドをJSONシリアライズ可能な構造に変換”””
rows_out = []
for row in grid:
if row.is_empty():
continue
cells_out = []
for cell in row.cells:
if cell.is_merged_sub:
continue
if cell.is_empty:
continue
c = {“value”: cell.text}
if cell.is_merged_main:
c[“span_col”] = cell.merge_span_col
c[“span_row”] = cell.merge_span_row
if cell.indent > 0:
c[“indent”] = cell.indent
if cell.is_bold:
c[“bold”] = True
cells_out.append(c)
if cells_out:
rows_out.append(cells_out)
return {“rows”: rows_out}

def sections_to_json(sections: list[Section]) -> list[dict]:
result = []
for sec in sections:
rows = trim_grid_rows(sec.rows)
if not rows:
continue
data = rows_to_table_data(rows, sec.indent_base)
data = [r for r in data if any(c.strip() for c in r)]
entry = {}
if sec.title:
entry[“title”] = sec.title
entry[“rows”] = data
result.append(entry)
return result

# ──────────────────────────────────────────────

# メイン変換処理

# ──────────────────────────────────────────────

def sheet_to_markdown(sheet, sheet_name: str = “”) -> str:
grid = build_grid(sheet)
sections = split_into_sections(grid)

```
parts = []
if sheet_name:
    parts.append(f"# {sheet_name}")
    parts.append("")

for sec in sections:
    md = section_to_markdown(sec, heading_level=2 if sheet_name else 1)
    if md.strip():
        parts.append(md)
        parts.append("")

return "\n".join(parts).rstrip()
```

def sheet_to_json(sheet) -> dict:
grid = build_grid(sheet)
sections = split_into_sections(grid)
return {
“sections”: sections_to_json(sections),
“raw”: grid_to_json(grid),
}

def workbook_to_markdown(wb, target_sheet: Optional[str] = None) -> str:
sheet_names = wb.sheetnames
if target_sheet:
sheet_names = [s for s in sheet_names if s == target_sheet]
if not sheet_names:
print(f”シート ‘{target_sheet}’ が見つかりません。”, file=sys.stderr)
sys.exit(1)

```
parts = []
multi = len(sheet_names) > 1
for name in sheet_names:
    ws = wb[name]
    md = sheet_to_markdown(ws, sheet_name=name if multi else "")
    if md.strip():
        parts.append(md)
        if multi:
            parts.append("\n---\n")

return "\n".join(parts).rstrip()
```

def workbook_to_json(wb, target_sheet: Optional[str] = None) -> dict:
sheet_names = wb.sheetnames
if target_sheet:
sheet_names = [s for s in sheet_names if s == target_sheet]

```
result = {}
for name in sheet_names:
    ws = wb[name]
    result[name] = sheet_to_json(ws)
return result
```

# ──────────────────────────────────────────────

# CLI

# ──────────────────────────────────────────────

def main():
parser = argparse.ArgumentParser(
description=“Excel方眼紙・管理会計PL → Markdown/JSON変換”
)
parser.add_argument(“input”, help=“入力 .xlsx ファイル”)
parser.add_argument(“output”, nargs=”?”, help=“出力ファイル（省略時は標準出力）”)
parser.add_argument(”–format”, choices=[“md”, “json”], default=“md”, help=“出力形式”)
parser.add_argument(”–sheet”, default=None, help=“対象シート名（省略時は全シート）”)
args = parser.parse_args()

```
wb = load_workbook(args.input, data_only=True)

if args.format == "md":
    output = workbook_to_markdown(wb, args.sheet)
else:
    data = workbook_to_json(wb, args.sheet)
    output = json.dumps(data, ensure_ascii=False, indent=2)

if args.output:
    Path(args.output).write_text(output, encoding="utf-8")
    print(f"変換完了: {args.output}")
else:
    print(output)
```

if **name** == “**main**”:
main()
